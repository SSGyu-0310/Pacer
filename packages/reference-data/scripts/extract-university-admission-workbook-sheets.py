#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

LOCAL_PYTHON_TOOLING = ".reference-data/tools/python"


def bootstrap_local_python_tooling() -> None:
    current = Path.cwd().resolve()
    for candidate_root in [current, *current.parents]:
        if (candidate_root / "pnpm-workspace.yaml").exists():
            tooling_path = candidate_root / LOCAL_PYTHON_TOOLING
            if tooling_path.exists():
                sys.path.insert(0, str(tooling_path))
            return


bootstrap_local_python_tooling()

from openpyxl import load_workbook


DEFAULT_MANIFESTS = [
    "packages/reference-data/data/public/university-admission-sites/university_admission_attachment_artifact_manifest_2027.jsonl",
    "packages/reference-data/data/public/university-admission-sites/university_admission_attachment_artifact_manifest_2027_file_download_route.jsonl",
    "packages/reference-data/data/public/university-admission-sites/university_admission_attachment_artifact_manifest_2027_related_detail.jsonl",
    "packages/reference-data/data/public/university-admission-sites/university_admission_attachment_artifact_manifest_2027_related_detail_file_routes.jsonl",
]
DEFAULT_OUTPUT_DIR = "packages/reference-data/data/public/university-admission-sites/extracted"
SUPPORTED_EXTENSIONS = {"xls", "xlsx"}
WORKBOOK_EXTENSIONS = {"xls", "xlsx"}
ADMISSION_DOMAIN_PATTERN = re.compile(
    r"모집|입시|입학|전형|정시|수시|수능|지원|경쟁률|합격|등록|충원|"
    r"모집단위|모집인원|성적|환산|백분위|등급|학생부|면접|실기|논술"
)
LIKELY_NON_ADMISSION_PATTERN = re.compile(
    r"testcard|visa|mastercard|credit\s*card|card\s*number|western\s+type|"
    r"check[-\s]?in|check[-\s]?out|passport|room|hotel|reservation|"
    r"wcc\d+|korea,\s*the\s+republic\s+of|@[a-z0-9_.+-]+|"
    r"호텔|숙박|예약번호|입실일|퇴실일|카드번호|카드소지자|룸타입|"
    r"합계금액|기본요금|전화번호|팩스|이메일",
    re.IGNORECASE,
)


def main() -> None:
    args = parse_args()
    repo_root = find_repo_root(Path.cwd())
    add_local_python_tooling(repo_root)
    xlrd_module = import_optional_xlrd()
    manifest_paths = [resolve(repo_root, value) for value in args.manifest]
    output_dir = resolve(repo_root, args.output_dir)
    csv_root = output_dir / "workbook-sheets" / str(args.year)
    csv_root.mkdir(parents=True, exist_ok=True)

    artifact_rows = load_manifest_rows(manifest_paths)
    workbook_rows = [
        row
        for row in artifact_rows
        if is_http_ok_file(row, repo_root)
        and int(row.get("year") or 0) == args.year
        and artifact_extension(row) in WORKBOOK_EXTENSIONS
    ]

    source_rows: list[dict[str, Any]] = []
    sheet_rows: list[dict[str, Any]] = []

    for index, artifact in enumerate(workbook_rows, start=1):
        source_row = source_manifest_row(artifact, repo_root)
        extension = source_row["fileExtension"]
        raw_workbook_path = repo_root / str(source_row["rawWorkbookPath"])

        if extension not in SUPPORTED_EXTENSIONS:
            source_row["extractionStatus"] = "unsupported_legacy_xls"
            source_row["notes"] = "Requires xlrd or a LibreOffice conversion pass before CSV extraction."
            source_rows.append(source_row)
            continue

        if extension == "xls" and xlrd_module is None:
            source_row["extractionStatus"] = "unsupported_legacy_xls"
            source_row["notes"] = (
                f"Install xlrd into {LOCAL_PYTHON_TOOLING} or use a LibreOffice conversion pass "
                "before CSV extraction."
            )
            source_rows.append(source_row)
            continue

        if not raw_workbook_path.exists():
            source_row["extractionStatus"] = "missing_raw_file"
            source_row["notes"] = "Raw workbook path does not exist in the local artifact store."
            source_rows.append(source_row)
            continue

        try:
            if extension == "xlsx":
                extracted_sheets = extract_xlsx_sheets(
                    source_row=source_row,
                    workbook_path=raw_workbook_path,
                    csv_root=csv_root,
                    repo_root=repo_root,
                    sequence=index,
                )
            else:
                extracted_sheets = extract_xls_sheets(
                    source_row=source_row,
                    workbook_path=raw_workbook_path,
                    csv_root=csv_root,
                    repo_root=repo_root,
                    sequence=index,
                    xlrd_module=xlrd_module,
                )
            source_row["extractionStatus"] = "extracted"
            source_row["sheetCount"] = len(extracted_sheets)
            source_row["notes"] = ""
            sheet_rows.extend(extracted_sheets)
        except Exception as error:
            source_row["extractionStatus"] = "extract_failed"
            source_row["notes"] = f"{type(error).__name__}: {error}"

        source_rows.append(source_row)

    write_jsonl(
        output_dir / f"university_admission_workbook_sources_manifest_{args.year}.jsonl",
        source_rows,
    )
    write_jsonl(
        output_dir / f"university_admission_workbook_sheets_manifest_{args.year}.jsonl",
        sheet_rows,
    )
    summary = summarize(args.year, source_rows, sheet_rows)
    (output_dir / f"university_admission_workbook_sheets_summary_{args.year}.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(
        "university admission workbook sheet extraction complete. "
        f"sources={summary['sourceWorkbooks']} "
        f"extractedWorkbooks={summary['extractedWorkbooks']} "
        f"sheets={summary['sheets']} "
        f"csvRoot={to_repo_relative(csv_root, repo_root)}"
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--year", type=int, default=2027)
    parser.add_argument("--manifest", action="append", default=[])
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args(cli_args())
    if not args.manifest:
        args.manifest = DEFAULT_MANIFESTS
    return args


def cli_args() -> list[str]:
    args = sys.argv[1:]
    return args[1:] if args[:1] == ["--"] else args


def find_repo_root(start: Path) -> Path:
    current = start.resolve()
    while True:
        if (current / "pnpm-workspace.yaml").exists():
            return current
        if current.parent == current:
            return start.resolve()
        current = current.parent


def resolve(repo_root: Path, value: str) -> Path:
    path = Path(value)
    return path if path.is_absolute() else repo_root / path


def add_local_python_tooling(repo_root: Path) -> None:
    tooling_path = repo_root / LOCAL_PYTHON_TOOLING
    if tooling_path.exists():
        sys.path.insert(0, str(tooling_path))


def import_optional_xlrd() -> Any | None:
    try:
        import xlrd  # type: ignore[import-not-found]

        return xlrd
    except Exception:
        return None


def load_manifest_rows(paths: list[Path]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for path in paths:
        for line in path.read_text(encoding="utf-8").split("\n"):
            if line.strip():
                row = json.loads(line)
                row["_manifestPath"] = str(path)
                rows.append(row)
    return rows


def is_http_ok_file(row: dict[str, Any], repo_root: Path) -> bool:
    status = str(row.get("status") or "")
    kind = str(row.get("detectedKind") or "")
    http_status = row.get("httpStatus")
    return (
        status == "fetched"
        and isinstance(http_status, int)
        and 200 <= http_status < 300
        and (kind == "file" or has_any_file_signature(row, repo_root, [b"\xd0\xcf\x11\xe0", b"PK"]))
    )


def artifact_extension(row: dict[str, Any]) -> str:
    extension = str(row.get("fileExtension") or "").lower().lstrip(".")
    if extension:
        return extension
    raw_suffix = Path(str(row.get("rawPath") or "")).suffix.lower().lstrip(".")
    if raw_suffix:
        return raw_suffix
    return Path(str(row.get("suggestedFilename") or "")).suffix.lower().lstrip(".")


def has_any_file_signature(row: dict[str, Any], repo_root: Path, signatures: list[bytes]) -> bool:
    raw_path = str(row.get("rawPath") or "")
    if not raw_path:
        return False
    path = repo_root / raw_path
    if not path.exists():
        return False
    max_len = max(len(signature) for signature in signatures)
    try:
        with path.open("rb") as file:
            prefix = file.read(max_len)
    except OSError:
        return False
    return any(prefix.startswith(signature) for signature in signatures)


def source_manifest_row(artifact: dict[str, Any], repo_root: Path) -> dict[str, Any]:
    raw_path = str(artifact.get("rawPath") or "")
    raw_workbook_path = repo_root / raw_path
    return {
        "provider": "university-admission-office",
        "artifactType": "admission_workbook_source",
        "year": artifact.get("year"),
        "unvCd": artifact.get("unvCd"),
        "universityName": artifact.get("universityName"),
        "campus": artifact.get("campus"),
        "sourceLinkRole": artifact.get("sourceLinkRole"),
        "attachmentRole": artifact.get("attachmentRole"),
        "linkText": artifact.get("linkText"),
        "sourceCandidateUrl": artifact.get("sourceCandidateUrl"),
        "attachmentUrl": artifact.get("attachmentUrl"),
        "finalUrl": artifact.get("finalUrl"),
        "rawWorkbookPath": raw_path,
        "rawWorkbookSha256": artifact.get("sha256"),
        "rawWorkbookBytes": artifact.get("bytes"),
        "fileExtension": artifact_extension(artifact),
        "suggestedFilename": artifact.get("suggestedFilename"),
        "contentType": artifact.get("contentType"),
        "sourceManifestPath": to_repo_relative(Path(str(artifact.get("_manifestPath"))), repo_root),
        "rawFileExists": raw_workbook_path.exists(),
        "sheetCount": 0,
        "extractedAt": datetime.now(timezone.utc).isoformat(),
    }


def extract_xlsx_sheets(
    *,
    source_row: dict[str, Any],
    workbook_path: Path,
    csv_root: Path,
    repo_root: Path,
    sequence: int,
) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_rows: list[dict[str, Any]] = []
    source_hash_prefix = str(source_row.get("rawWorkbookSha256") or "unknown")[:16] or "unknown"
    workbook_dir = (
        csv_root
        / str(source_row.get("unvCd") or "unknown")
        / f"{sequence:04d}_{source_hash_prefix}"
    )
    workbook_dir.mkdir(parents=True, exist_ok=True)

    for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
        rows = list(sheet.iter_rows(values_only=True))
        trimmed_rows = trim_rows(rows)
        sheet_rows.append(
            make_sheet_manifest_row(
                source_row=source_row,
                sheet_name=sheet.title,
                trimmed_rows=trimmed_rows,
                csv_path=workbook_dir / f"{sheet_index:02d}_{safe_filename(sheet.title)}.csv",
                repo_root=repo_root,
            )
        )

    return sheet_rows


def extract_xls_sheets(
    *,
    source_row: dict[str, Any],
    workbook_path: Path,
    csv_root: Path,
    repo_root: Path,
    sequence: int,
    xlrd_module: Any,
) -> list[dict[str, Any]]:
    workbook = xlrd_module.open_workbook(str(workbook_path), on_demand=True)
    sheet_rows: list[dict[str, Any]] = []
    source_hash_prefix = str(source_row.get("rawWorkbookSha256") or "unknown")[:16] or "unknown"
    workbook_dir = (
        csv_root
        / str(source_row.get("unvCd") or "unknown")
        / f"{sequence:04d}_{source_hash_prefix}"
    )
    workbook_dir.mkdir(parents=True, exist_ok=True)

    for sheet_index, sheet in enumerate(workbook.sheets(), start=1):
        rows = [
            [
                normalize_xls_cell(
                    sheet.cell(row_index, col_index),
                    datemode=workbook.datemode,
                    xlrd_module=xlrd_module,
                )
                for col_index in range(sheet.ncols)
            ]
            for row_index in range(sheet.nrows)
        ]
        trimmed_rows = trim_rows(rows)
        sheet_rows.append(
            make_sheet_manifest_row(
                source_row=source_row,
                sheet_name=sheet.name,
                trimmed_rows=trimmed_rows,
                csv_path=workbook_dir / f"{sheet_index:02d}_{safe_filename(sheet.name)}.csv",
                repo_root=repo_root,
            )
        )

    workbook.release_resources()
    return sheet_rows


def make_sheet_manifest_row(
    *,
    source_row: dict[str, Any],
    sheet_name: str,
    trimmed_rows: list[list[Any]],
    csv_path: Path,
    repo_root: Path,
) -> dict[str, Any]:
    header_preview = preview_rows(trimmed_rows)
    role_text = " ".join(
        [
            str(source_row.get("sourceLinkRole") or ""),
            str(source_row.get("linkText") or ""),
            sheet_name,
            flatten_preview(header_preview),
        ]
    )
    detected_sheet_role = detect_sheet_role(role_text)
    excluded_non_admission = is_likely_non_admission_sheet(trimmed_rows)
    if not excluded_non_admission:
        write_csv(csv_path, trimmed_rows)

    return {
        "provider": "university-admission-office",
        "artifactType": "admission_workbook_sheet_csv",
        "year": source_row.get("year"),
        "unvCd": source_row.get("unvCd"),
        "universityName": source_row.get("universityName"),
        "campus": source_row.get("campus"),
        "sourceLinkRole": source_row.get("sourceLinkRole"),
        "attachmentRole": source_row.get("attachmentRole"),
        "linkText": source_row.get("linkText"),
        "sheetName": sheet_name,
        "detectedSheetRole": detected_sheet_role,
        "sourceCandidateUrl": source_row.get("sourceCandidateUrl"),
        "attachmentUrl": source_row.get("attachmentUrl"),
        "finalUrl": source_row.get("finalUrl"),
        "rawWorkbookPath": source_row.get("rawWorkbookPath"),
        "rawWorkbookSha256": source_row.get("rawWorkbookSha256"),
        "fileExtension": source_row.get("fileExtension"),
        "csvPath": "" if excluded_non_admission else to_repo_relative(csv_path, repo_root),
        "rows": len(trimmed_rows),
        "cols": max((len(row) for row in trimmed_rows), default=0),
        "nonEmptyCells": count_non_empty_cells(trimmed_rows),
        "headerPreview": [] if excluded_non_admission else header_preview,
        "sha256": "" if excluded_non_admission else sha256_file(csv_path),
        "extractedAt": datetime.now(timezone.utc).isoformat(),
        "status": "excluded_non_admission_sheet" if excluded_non_admission else "extracted",
        "notes": (
            "Excluded from public sheet CSV export because the workbook content looks unrelated "
            "to admission data and contains personal/payment-style fields."
            if excluded_non_admission
            else ""
        ),
    }


def trim_rows(rows: list[tuple[Any, ...]]) -> list[list[Any]]:
    last_row = -1
    last_col = -1
    for row_index, row in enumerate(rows):
        for col_index, value in enumerate(row):
            if value not in (None, ""):
                last_row = max(last_row, row_index)
                last_col = max(last_col, col_index)

    if last_row < 0 or last_col < 0:
        return []

    return [
        [normalize_cell(value) for value in row[: last_col + 1]]
        for row in rows[: last_row + 1]
    ]


def normalize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value).strip()
    return value


def normalize_xls_cell(cell: Any, *, datemode: int, xlrd_module: Any) -> Any:
    if cell.ctype == xlrd_module.XL_CELL_EMPTY:
        return ""
    if cell.ctype == xlrd_module.XL_CELL_DATE:
        try:
            return xlrd_module.xldate.xldate_as_datetime(cell.value, datemode).isoformat(sep=" ")
        except Exception:
            return cell.value
    if cell.ctype == xlrd_module.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd_module.XL_CELL_ERROR:
        return f"#ERROR({cell.value})"
    if cell.ctype == xlrd_module.XL_CELL_NUMBER:
        number = float(cell.value)
        return int(number) if number.is_integer() else number
    return normalize_cell(cell.value)


def write_csv(path: Path, rows: list[list[Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerows(rows)


def count_non_empty_cells(rows: list[list[Any]]) -> int:
    return sum(1 for row in rows for value in row if value not in ("", None))


def preview_rows(rows: list[list[Any]]) -> list[list[str]]:
    return [[str(value) for value in row[:10]] for row in rows[:6]]


def flatten_preview(rows: list[list[str]]) -> str:
    return " ".join(value for row in rows for value in row)


def is_likely_non_admission_sheet(rows: list[list[Any]]) -> bool:
    flattened = " ".join(str(value) for row in rows[:20] for value in row if value not in ("", None))
    if not flattened:
        return False
    admission_hits = len(ADMISSION_DOMAIN_PATTERN.findall(flattened))
    non_admission_hits = len(LIKELY_NON_ADMISSION_PATTERN.findall(flattened))
    has_email = bool(re.search(r"[\w.+-]+@[\w.-]+\.[a-z]{2,}", flattened, re.IGNORECASE))
    has_card_like_number = bool(re.search(r"\b(?:\d[ -]?){13,19}\b", flattened))
    return (non_admission_hits >= 3 and admission_hits <= 2) or (
        has_email and has_card_like_number and admission_hits <= 2
    )


def detect_sheet_role(text: str) -> str:
    normalized = text.lower()
    if re.search(r"경쟁률|competition", normalized):
        return "competition_rate_table"
    if re.search(r"입시결과|입학결과|전형결과|합격|등록|충원|성적|환산|백분위|등급", normalized):
        return "admission_result_table"
    if re.search(r"모집요강|모집인원|전형계획|정시|수능|전형방법|반영", normalized):
        return "recruitment_notice_table"
    return "unknown"


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^\w.()\-가-힣]+", "_", value)
    cleaned = re.sub(r"_+", "_", cleaned)
    return cleaned.strip("_")[:120] or "sheet"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def to_repo_relative(path: Path, repo_root: Path) -> str:
    resolved = path.resolve()
    try:
        return str(resolved.relative_to(repo_root))
    except ValueError:
        return str(resolved)


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "\n".join(json.dumps(sanitize_json_value(row), ensure_ascii=False) for row in rows) + "\n",
        encoding="utf-8",
    )


def sanitize_json_value(value: Any) -> Any:
    if isinstance(value, str):
        return re.sub(r"[\u0000-\u001f\u007f-\u009f]+", " ", value).strip()
    if isinstance(value, list):
        return [sanitize_json_value(item) for item in value]
    if isinstance(value, dict):
        return {key: sanitize_json_value(item) for key, item in value.items()}
    return value


def count_by(rows: list[dict[str, Any]], key: str) -> list[dict[str, Any]]:
    counts: dict[str, int] = {}
    for row in rows:
        value = str(row.get(key) or "")
        counts[value] = counts.get(value, 0) + 1
    return [
        {"value": value, "count": count}
        for value, count in sorted(counts.items(), key=lambda item: (-item[1], item[0]))
    ]


def summarize(
    year: int,
    source_rows: list[dict[str, Any]],
    sheet_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    extracted_sources = [
        row for row in source_rows if row.get("extractionStatus") == "extracted"
    ]
    return {
        "provider": "university-admission-office",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "year": year,
        "sourceWorkbooks": len(source_rows),
        "extractedWorkbooks": len(extracted_sources),
        "unsupportedLegacyXlsWorkbooks": sum(
            1 for row in source_rows if row.get("extractionStatus") == "unsupported_legacy_xls"
        ),
        "failedWorkbooks": sum(
            1 for row in source_rows if row.get("extractionStatus") == "extract_failed"
        ),
        "sheets": len(sheet_rows),
        "extractedSheets": sum(1 for row in sheet_rows if row.get("status") == "extracted"),
        "excludedNonAdmissionSheets": sum(
            1 for row in sheet_rows if row.get("status") == "excluded_non_admission_sheet"
        ),
        "totalSheetRows": sum(
            int(row.get("rows") or 0) for row in sheet_rows if row.get("status") == "extracted"
        ),
        "totalNonEmptyCells": sum(
            int(row.get("nonEmptyCells") or 0)
            for row in sheet_rows
            if row.get("status") == "extracted"
        ),
        "uniqueRawWorkbookSha256": len(
            {str(row.get("rawWorkbookSha256") or "") for row in source_rows}
        ),
        "extractedXlsxWorkbooks": sum(
            1
            for row in extracted_sources
            if str(row.get("fileExtension") or "").lower() == "xlsx"
        ),
        "extractedLegacyXlsWorkbooks": sum(
            1 for row in extracted_sources if str(row.get("fileExtension") or "").lower() == "xls"
        ),
        "bySourceExtension": count_by(source_rows, "fileExtension"),
        "byExtractionStatus": count_by(source_rows, "extractionStatus"),
        "bySheetStatus": count_by(sheet_rows, "status"),
        "bySourceLinkRole": count_by(source_rows, "sourceLinkRole"),
        "byDetectedSheetRole": count_by(sheet_rows, "detectedSheetRole"),
        "notes": [
            "Workbook sources come from downloaded university admission-office attachment manifests.",
            "XLSX files are extracted with openpyxl; legacy XLS files are extracted with xlrd when available.",
            "CSV sheets are source-preserving candidates and require human verification before promotion to AdmissionRule or HistoricalOutcome.",
        ],
    }


if __name__ == "__main__":
    main()
