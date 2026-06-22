#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
import zipfile
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

LOCAL_PYTHON_TOOLING = ".reference-data/tools/python"


def bootstrap_local_python_tooling() -> None:
    current = Path.cwd().resolve()
    for candidate_root in [current, *current.parents]:
        if (candidate_root / "pnpm-workspace.yaml").exists():
            tooling_path = candidate_root / LOCAL_PYTHON_TOOLING
            if tooling_path.exists():
                sys.path.insert(0, str(tooling_path))
            return


bootstrap_local_python_tooling()

from openpyxl import load_workbook


DEFAULT_MANIFEST = (
    "packages/reference-data/data/public/academyinfo/"
    "academyinfo_public_data_download_manifest.jsonl"
)
DEFAULT_OUTPUT_DIR = "packages/reference-data/data/public/academyinfo/extracted"


def main() -> None:
    args = parse_args()
    repo_root = find_repo_root(Path.cwd())
    manifest_path = resolve(repo_root, args.manifest)
    output_dir = resolve(repo_root, args.output_dir)
    csv_root = output_dir / "workbook-sheets"
    csv_root.mkdir(parents=True, exist_ok=True)

    manifest_rows = load_jsonl(manifest_path)
    source_rows: list[dict[str, Any]] = []
    sheet_rows: list[dict[str, Any]] = []

    for sequence, row in enumerate(manifest_rows, start=1):
        if row.get("status") != "downloaded" or not row.get("rawZipPath"):
            continue

        source_row = base_source_row(row, sequence, repo_root)
        zip_path = repo_root / str(row["rawZipPath"])
        if not zip_path.exists():
            source_row["extractionStatus"] = "missing_zip"
            source_row["notes"] = "rawZipPath does not exist"
            source_rows.append(source_row)
            continue

        try:
            extracted_sheets = extract_zip_workbooks(
                source_row=source_row,
                zip_path=zip_path,
                csv_root=csv_root,
                repo_root=repo_root,
            )
            source_row["extractionStatus"] = "extracted"
            source_row["workbookCount"] = len({sheet["innerWorkbookName"] for sheet in extracted_sheets})
            source_row["sheetCount"] = len(extracted_sheets)
            source_row["notes"] = ""
            sheet_rows.extend(extracted_sheets)
        except Exception as error:
            source_row["extractionStatus"] = "extract_failed"
            source_row["notes"] = f"{type(error).__name__}: {error}"

        source_rows.append(source_row)

    write_jsonl(output_dir / "academyinfo_workbook_sources_manifest.jsonl", source_rows)
    write_jsonl(output_dir / "academyinfo_workbook_sheets_manifest.jsonl", sheet_rows)
    write_csv(
        output_dir / "academyinfo_workbook_sheets_index.csv",
        sheet_rows,
        [
            "surveyYear",
            "itemId",
            "relevanceRole",
            "outputKindCode",
            "outputKindLabel",
            "innerWorkbookName",
            "sheetName",
            "rows",
            "cols",
            "nonEmptyCells",
            "csvPath",
            "sourceZipPath",
        ],
    )
    summary = summarize(source_rows, sheet_rows)
    (output_dir / "academyinfo_workbook_sheets_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(
        "academyinfo workbook extraction complete. "
        f"sources={summary['sources']} workbooks={summary['workbooks']} "
        f"sheets={summary['sheets']} csvRoot={to_repo_relative(csv_root, repo_root)}"
    )


def extract_zip_workbooks(
    source_row: dict[str, Any],
    zip_path: Path,
    csv_root: Path,
    repo_root: Path,
) -> list[dict[str, Any]]:
    sheet_rows: list[dict[str, Any]] = []
    with zipfile.ZipFile(zip_path) as archive:
        workbook_names = [
            name
            for name in archive.namelist()
            if not name.endswith("/") and name.lower().endswith(".xlsx")
        ]
        if not workbook_names:
            raise ValueError("no xlsx workbook found in zip")

        for workbook_index, workbook_name in enumerate(workbook_names, start=1):
            workbook_bytes = archive.read(workbook_name)
            workbook_sha = sha256_bytes(workbook_bytes)
            workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
            workbook_rel_dir = Path(
                str(source_row["surveyYear"]),
                str(source_row["itemId"]),
                str(source_row["outputKindCode"]),
                f"{source_row['sequence']:04d}_{workbook_index:02d}_{workbook_sha[:16]}",
            )

            for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
                sheet.reset_dimensions()
                rows = list(sheet.iter_rows(values_only=True))
                trimmed_rows = trim_rows(rows)
                csv_dir = csv_root / workbook_rel_dir
                csv_dir.mkdir(parents=True, exist_ok=True)
                csv_path = csv_dir / f"{sheet_index:02d}_{safe_filename(sheet.title)}.csv"
                write_csv_rows(csv_path, trimmed_rows)

                sheet_rows.append(
                    {
                        "provider": "academyinfo",
                        "artifactType": "academyinfo_workbook_sheet_csv",
                        "sequence": source_row["sequence"],
                        "surveyYear": source_row["surveyYear"],
                        "itemId": source_row["itemId"],
                        "itemDivCd": source_row["itemDivCd"],
                        "relevanceRole": source_row["relevanceRole"],
                        "pacerTargets": source_row["pacerTargets"],
                        "outputKindCode": source_row["outputKindCode"],
                        "outputKindLabel": source_row["outputKindLabel"],
                        "academyinfoFileName": source_row["academyinfoFileName"],
                        "innerWorkbookName": workbook_name,
                        "innerWorkbookIndex": workbook_index,
                        "innerWorkbookSha256": workbook_sha,
                        "sheetName": sheet.title,
                        "sheetIndex": sheet_index,
                        "sourceZipPath": source_row["sourceZipPath"],
                        "csvPath": to_repo_relative(csv_path, repo_root),
                        "rows": len(trimmed_rows),
                        "cols": max((len(row) for row in trimmed_rows), default=0),
                        "nonEmptyCells": count_non_empty_cells(trimmed_rows),
                        "headerPreview": preview_rows(trimmed_rows),
                        "sha256": sha256_file(csv_path),
                        "extractedAt": datetime.now(timezone.utc).isoformat(),
                        "status": "extracted",
                    }
                )

    return sheet_rows


def base_source_row(row: dict[str, Any], sequence: int, repo_root: Path) -> dict[str, Any]:
    zip_path = repo_root / str(row["rawZipPath"])
    return {
        "provider": "academyinfo",
        "artifactType": "academyinfo_workbook_source",
        "sequence": sequence,
        "surveyYear": row.get("surveyYear"),
        "itemId": row.get("itemId"),
        "itemDivCd": row.get("itemDivCd"),
        "relevanceRole": row.get("relevanceRole"),
        "pacerTargets": row.get("pacerTargets"),
        "schoolDivisionCode": row.get("schoolDivisionCode"),
        "outputKindCode": row.get("outputKindCode"),
        "outputKindLabel": row.get("outputKindLabel"),
        "academyinfoPath": row.get("academyinfoPath"),
        "academyinfoFileName": row.get("academyinfoFileName"),
        "sourceZipPath": row.get("rawZipPath"),
        "sourceZipSha256": row.get("sha256"),
        "sourceZipBytes": row.get("bytes"),
        "sourceZipActualSha256": sha256_file(zip_path) if zip_path.exists() else None,
        "sourceZipActualBytes": zip_path.stat().st_size if zip_path.exists() else None,
        "workbookCount": 0,
        "sheetCount": 0,
        "extractionStatus": "pending",
        "notes": "",
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--manifest", default=DEFAULT_MANIFEST)
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args(cli_args())


def cli_args() -> list[str]:
    args = sys.argv[1:]
    return args[1:] if args[:1] == ["--"] else args


def find_repo_root(start: Path) -> Path:
    current = start.resolve()
    while True:
        if (current / "pnpm-workspace.yaml").exists():
            return current
        if current.parent == current:
            return start.resolve()
        current = current.parent


def resolve(repo_root: Path, value: str) -> Path:
    path = Path(value)
    return path if path.is_absolute() else repo_root / path


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]


def trim_rows(rows: list[tuple[Any, ...]]) -> list[list[Any]]:
    matrix = [list(row) for row in rows]
    while matrix and all(is_blank(cell) for cell in matrix[-1]):
        matrix.pop()
    while matrix and all(is_blank(row[0] if row else None) for row in matrix):
        matrix = [row[1:] for row in matrix]
    while matrix and all(is_blank(row[-1] if row else None) for row in matrix):
        matrix = [row[:-1] for row in matrix]
    return matrix


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def write_csv_rows(path: Path, rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    column: "|".join(map(str, row[column]))
                    if isinstance(row.get(column), list)
                    else row.get(column)
                    for column in columns
                }
            )


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows),
        encoding="utf-8",
    )


def count_non_empty_cells(rows: list[list[Any]]) -> int:
    return sum(1 for row in rows for cell in row if not is_blank(cell))


def preview_rows(rows: list[list[Any]], limit: int = 8) -> list[str]:
    return [
        " | ".join("" if cell is None else str(cell) for cell in row)
        for row in rows[:limit]
    ]


def summarize(source_rows: list[dict[str, Any]], sheet_rows: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "provider": "academyinfo",
        "sources": len(source_rows),
        "sourcesExtracted": sum(1 for row in source_rows if row.get("extractionStatus") == "extracted"),
        "sourcesFailed": sum(1 for row in source_rows if row.get("extractionStatus") == "extract_failed"),
        "workbooks": len({row["innerWorkbookSha256"] for row in sheet_rows}),
        "sheets": len(sheet_rows),
        "rows": sum(int(row.get("rows") or 0) for row in sheet_rows),
        "nonEmptyCells": sum(int(row.get("nonEmptyCells") or 0) for row in sheet_rows),
        "roles": count_by(sheet_rows, "relevanceRole"),
        "years": count_by(sheet_rows, "surveyYear"),
        "outputKinds": count_by(sheet_rows, "outputKindLabel"),
        "generatedAt": datetime.now(timezone.utc).isoformat(),
    }


def count_by(rows: list[dict[str, Any]], key: str) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        value = str(row.get(key))
        counts[value] = counts.get(value, 0) + 1
    return counts


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^\w.\-가-힣]+", "_", value).strip("_")
    return cleaned or "sheet"


def to_repo_relative(path: Path, repo_root: Path) -> str:
    return path.relative_to(repo_root).as_posix()


if __name__ == "__main__":
    main()
